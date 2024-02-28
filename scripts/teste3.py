# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

translate = {
    # ... (seu dicionário)
}

def traduzir_colunas(coluna, translate):
    return ' '.join(translate.get(palavra, palavra) for palavra in re.findall(r'\b\w+\b', coluna)).upper()

def carregar_excel(caminho_excel):
    workbook = load_workbook(caminho_excel)
    sheet = workbook.active
    data = sheet.values
    columns = next(data)[:5]
    return list(data), columns

def salvar_excel(caminho_saida_excel, data, columns):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(columns)
    for row in data:
        sheet.append(row)
    workbook.save(caminho_saida_excel)

def main():
    caminho_excel = '/var/www/html/python2.7_translater_files_geoprocessing/files/for_test_setor.xlsx'
    data, columns = carregar_excel(caminho_excel)

    # translate
    colunas_a_traduzir = ['se_setor', 'qu_setor', 'nome']
    
    for coluna_index, coluna_nome in enumerate(columns):
        if coluna_nome in colunas_a_traduzir:
            coluna = [row[coluna_index] for row in data]
            data = [row[:coluna_index] + (traduzir_colunas(palavra, translate),) + row[coluna_index + 1:] for row, palavra in zip(data, coluna)]

    caminho_saida_excel = '/var/www/html/python2.7_translater_files_geoprocessing/files/resultados/traducao_results/translate_table.xlsx'
    salvar_excel(caminho_saida_excel, data, columns)

if __name__ == "__main__":
    main()
